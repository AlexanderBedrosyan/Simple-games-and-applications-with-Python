from tkinter import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

root = Tk()
root.title("Excell compare.app")
root.geometry("1000x500")
root.config(bg="#EEEEEE")
root.resizable(False, False)

file_one_dict = {}
file_two_dict = {}


def check():
    name_one = entry_file_first.get()
    name_two = entry_file_second.get()
    first_column = entry_column_one.get()
    second_column = entry_column_two.get()

    wb = Workbook()
    wb = load_workbook(name_one + '.xlsx')
    ws = wb.active
    column_a = ws[first_column.upper()]
    column_b = ws[second_column.upper()]
    list_column_a = []
    list_column_b = []
    for cell in column_a:
        list_column_a.append(cell.value)
    for cell in column_b:
        list_column_b.append(cell.value)
    for i in range(len(list_column_a)):
        file_one_dict[list_column_a[i]] = list_column_b[i]

    wb = Workbook()
    wb = load_workbook(name_two + '.xlsx')
    ws = wb.active
    column_a = ws[first_column.upper()]
    column_b = ws[second_column.upper()]
    list_column_a = []
    list_column_b = []
    for cell in column_a:
        list_column_a.append(cell.value)
    for cell in column_b:
        list_column_b.append(cell.value)
    for i in range(len(list_column_a)):
        file_two_dict[list_column_a[i]] = list_column_b[i]

    missing_documents_first_file = {}
    wrong_invoice_amount = {}
    missing_documents_second_file = {}
    for ch in file_one_dict:
        if ch not in file_two_dict:
            missing_documents_second_file[ch] = file_one_dict[ch]
        else:
            for item in file_two_dict:
                if ch == item:
                    if file_two_dict[item] == file_one_dict[item]:
                        continue
                    else:
                        wrong_invoice_amount[ch] = file_one_dict[ch]

    for ch in file_two_dict:
        if ch not in file_one_dict:
            missing_documents_first_file[ch] = file_two_dict[ch]

    result_window = Toplevel(root)
    result_window.title("Result")
    result_text = Text(result_window)
    result_text.insert(END, f"First file missing invoices:{missing_documents_first_file}\n"
                            f"Second file missing invoices:{missing_documents_second_file}\n"
                            f"Wrong invoice amount:{wrong_invoice_amount}")
    result_text.pack()


file_one = Label(root, text="First file name", font=4)
file_one.place(y=20)
entry_file_first = Entry(root, font=4)
entry_file_first.place(x=3, y=60)

file_two = Label(root, text="Second file name", font=4)
file_two.place(y=100)
entry_file_second = Entry(root, font=4)
entry_file_second.place(x=3, y=140)

column_one = Label(root, text="First column needed", font=4)
column_one.place(y=180)
entry_column_one = Entry(root, font=4)
entry_column_one.place(x=3, y=220)

column_two = Label(root, text="Second column needed", font=4)
column_two.place(y=260)
entry_column_two = Entry(root, font=4)
entry_column_two.place(x=3, y=300)

checking_button = Button(root, text="Check for differences", font=3, command=check)
checking_button.place(y=340)

root.mainloop()
